from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.models import Report


def build_report_workbook(report: Report) -> BytesIO:
    return _build_workbook([report])


def build_batch_workbook(reports: list[Report]) -> BytesIO:
    return _build_workbook(reports, include_network=True)


def _build_workbook(reports: list[Report], include_network: bool = False) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Распространение"

    column_count = 12 if include_network else 11
    last_column = get_column_letter(column_count)
    placement_start = "H" if include_network else "G"

    ws.merge_cells(f"A1:{last_column}1")
    ws["A1"] = "Распространение на медиаресурсах АО"
    ws.merge_cells("A2:F2")
    ws["A2"] = "Инфоповод"
    ws.merge_cells(f"{placement_start}2:{last_column}2")
    ws[f"{placement_start}2"] = "Размещение"

    headers = [
        "№ п/п",
        "АО",
        "ИО",
        "Дата",
        "Наименование инфоповода",
        "Ссылка на первоисточник",
        "Категория соц. сетей (окружное или районное сообщество/личная страница ЛОМа/другое(указать категорию))",
        "Наименование канала/сообщества",
        "Ссылка на публикацию",
        "Кол-во подписчиков",
        "Кол-во просмотров",
    ]
    if include_network:
        headers.insert(6, "Район")
    for col, value in enumerate(headers, start=1):
        ws.cell(row=3, column=col, value=value)

    row_idx = 4
    row_number = 1
    for report in reports:
        for row in report.rows:
            values = [
                row_number,
                "ЮВАО",
                "198",
                row.repost_date.strftime("%d.%m.%Y"),
                report.infopovod_title,
                report.source_url,
                row.resource.category_name,
                row.resource.display_name,
                row.repost_url,
                row.followers_count,
                row.views_count,
            ]
            if include_network:
                values.insert(6, report.network.name)
            for col, value in enumerate(values, start=1):
                ws.cell(row=row_idx, column=col, value=value)
            row_number += 1
            row_idx += 1

    pink = PatternFill("solid", fgColor="F4CCCC")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    for row in ws.iter_rows(
        min_row=1, max_row=max(ws.max_row, 3), min_col=1, max_col=column_count
    ):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if cell.row <= 3:
                cell.fill = pink
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    widths = [8, 10, 8, 14, 32, 36, 48, 32, 36, 18, 18]
    if include_network:
        widths.insert(6, 20)
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
