import argparse
import asyncio
from pathlib import Path
from urllib.parse import urlparse

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.exc import IntegrityError

from app.db import SessionLocal
from app.models import Network, Resource, ResourceType
from app.services.vk_client import VkClient

SHEET_MAP = {
    "Соц. сети управ и префектур": {
        "district_col": 3,
        "name_col": 5,
        "vk_col": 8,
        "resource_type": ResourceType.district_community,
        "category_label": None,
    },
    "Собств. неофиц. ресурсы": {
        "district_col": 3,
        "name_col": 5,
        "vk_col": 8,
        "resource_type": ResourceType.other,
        "category_label": "иной ресурс",
    },
    "Партнерские ресурсы": {
        "district_col": 3,
        "name_col": 5,
        "vk_col": 9,
        "resource_type": ResourceType.other,
        "category_label": "иной ресурс",
    },
    "Стр. сотруд. управ и префектур": {
        "district_col": 3,
        "name_col": 5,
        "vk_col": 9,
        "resource_type": ResourceType.lom_personal,
        "category_label": None,
    },
    "Личные страницы мун. депутатов": {
        "district_col": 3,
        "name_col": 5,
        "vk_col": 8,
        "resource_type": ResourceType.lom_personal,
        "category_label": None,
    },
    "Личные страницы ЛОМов": {
        "district_col": 3,
        "name_col": 5,
        "vk_col": 9,
        "resource_type": ResourceType.lom_personal,
        "category_label": None,
    },
}


def extract_screen_name(value: object) -> str | None:
    if value is None:
        return None
    raw = str(value).strip()
    if not raw or raw == "-":
        return None
    if raw.startswith("@"):
        return raw[1:].strip()
    if raw.startswith("http://") or raw.startswith("https://"):
        parsed = urlparse(raw)
        parts = [part for part in parsed.path.split("/") if part]
        return parts[0].strip() if parts else None
    if raw.startswith("vk.com/"):
        return raw.split("/", 1)[1].strip()
    return raw.strip()


def read_items(
    path: Path, allowed_districts: set[str] | None = None
) -> list[dict[str, str | ResourceType | None]]:
    wb = load_workbook(path, data_only=True, read_only=True)
    items: list[dict[str, str | ResourceType | None]] = []
    seen: set[tuple[str, str]] = set()
    for sheet_name, spec in SHEET_MAP.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=4, values_only=True):
            district = row[spec["district_col"] - 1] if len(row) >= spec["district_col"] else None
            district = str(district).strip() if district else None
            if allowed_districts is not None and district not in allowed_districts:
                continue
            name = row[spec["name_col"] - 1] if len(row) >= spec["name_col"] else None
            screen_name = extract_screen_name(row[spec["vk_col"] - 1] if len(row) >= spec["vk_col"] else None)
            if not name or not screen_name:
                continue
            dedupe_key = (district or "", screen_name.lower())
            if dedupe_key in seen:
                continue
            seen.add(dedupe_key)
            items.append(
                {
                    "district": district,
                    "display_name": str(name).strip(),
                    "screen_name": screen_name,
                    "resource_type": spec["resource_type"],
                    "category_label": spec["category_label"],
                }
            )
    return items


async def resolve_owner_id(client: VkClient, screen_name: str) -> int:
    resolved = await client.resolve_screen_name(screen_name)
    object_id = int(resolved["object_id"])
    return -object_id if resolved["type"] == "group" else object_id


async def import_resources(
    path: Path, network_name: str | None, district_names: list[str], dry_run: bool
) -> None:
    if network_name:
        district_names = [network_name]
    allowed_districts = set(district_names) if district_names else None
    items = read_items(path, allowed_districts=allowed_districts)
    print(f"Found {len(items)} VK resources in {path.name}")

    client = VkClient()
    created = 0
    updated = 0
    failed = 0
    with SessionLocal() as db:
        for item in items:
            district = str(item["district"] or network_name or "")
            network = db.scalar(select(Network).where(Network.name == district))
            if not network:
                failed += 1
                print(f"FAIL {item['screen_name']}: network not found: {district}")
                continue
            screen_name = str(item["screen_name"])
            try:
                owner_id = await resolve_owner_id(client, screen_name)
            except Exception as exc:
                failed += 1
                print(f"FAIL {screen_name}: {exc}")
                continue

            existing = db.scalar(
                select(Resource).where(
                    Resource.network_id == network.id,
                    Resource.vk_owner_id == owner_id,
                )
            )
            if existing:
                existing.vk_screen_name = screen_name
                existing.display_name = str(item["display_name"])
                existing.resource_type = item["resource_type"]
                existing.category_label = item["category_label"]
                existing.is_active = True
                updated += 1
                print(f"UPDATE {district}: {screen_name} -> {owner_id}")
            else:
                db.add(
                    Resource(
                        network_id=network.id,
                        vk_owner_id=owner_id,
                        vk_screen_name=screen_name,
                        display_name=str(item["display_name"]),
                        resource_type=item["resource_type"],
                        category_label=item["category_label"],
                        is_active=True,
                    )
                )
                created += 1
                print(f"CREATE {district}: {screen_name} -> {owner_id}")

            if not dry_run:
                try:
                    db.commit()
                except IntegrityError as exc:
                    db.rollback()
                    failed += 1
                    print(f"FAIL {screen_name}: {exc}")
            else:
                db.rollback()

            await asyncio.sleep(0.34)

    print(f"Done: created={created}, updated={updated}, failed={failed}, dry_run={dry_run}")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("path", type=Path)
    parser.add_argument("--network")
    parser.add_argument("--district", action="append", default=[])
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()
    asyncio.run(import_resources(args.path, args.network, args.district, args.dry_run))


if __name__ == "__main__":
    main()
